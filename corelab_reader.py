# -*- coding: utf-8 -*-
"""
Created on Mon Jun  8 19:26:05 2020

@author: Dmitry Molokhov (molokhov@outlook.com)
"""

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re


# Class to store sample data and relevant fluid properties.
class FlashExperimentData:
    
    def __init__(self, lqd=None, gas=None, res=None, ave_lqd_mw=None):
        self.liquid = lqd
        self.gas = gas
        self.reservoir = res
        self.av_lqd_mw = ave_lqd_mw
        # Estimation of average MW of C7 & C10 plus fractions
        # can be implemented later.
        self._ave_C10_mw = None
        self._ave_C7_mw = None
        self._c10_heavy_end_lqd = None
        self._c7_heavy_end_lqd = None
        self.gamma_input = None
    
    def assign_composition(self):
        pass
    
    # The following property decorators are to define the heavy 
    # end of the liquid fraction.
    # Don't like this way of splitting the heavy end but I reckon
    # I exhausted my knowledge of Pandas.
    @property
    def c10_heavy_end_lqd(self):
        i = self.liquid[self.liquid['scn'] == 'C10'].index[0]
        self._c10_heavy_end_lqd = self.liquid.iloc[i:]
        return self._c10_heavy_end_lqd

    @property
    def ave_C10_mw(self):
        self._ave_C10_mw = 225
        return self._ave_C10_mw
    
    @property
    def c7_heavy_end_lqd(self):
        i = self.liquid[self.liquid['scn'] == 'C7'].index[0]
        self._c7_heavy_end_lqd = self.liquid.iloc[i:]
        return self._c7_heavy_end_lqd
    
    def prepare_input(self, id, n=10):
        lqd=self.c10_heavy_end_lqd
        mw=self.av_lqd_mw
        self.gamma_input = lqd.copy()
        self.gamma_input['MWi_lab'] = self.gamma_input.apply(lambda x : x['lqd_wp']*mw/x['lqd_mp'], axis = 1)
        # Calculating initial values of upper bounds of individual MWn slices.
        # Upper bounds are considered as midway between SCN MW values.
        # C36+ upper bound is set to an arbitrary number (10000).
        self.gamma_input['ubound_init'] = self.gamma_input['MWi_lab']+(self.gamma_input['MWi_lab'].shift(-1)-self.gamma_input['MWi_lab'])/2
        self.gamma_input['ubound_init'] = self.gamma_input['ubound_init'].fillna(100000)
        
        # Generating regresion variables for component molecular weight bounds:
        self.gamma_input['ubound'] = 'm'+self.gamma_input['scn']
        
        # Adding top row to represent a lower boundary of C10 (or upper C9 boundary)
        top_index = 'C'+str(n-1)
        df_top = pd.DataFrame(pd.DataFrame([[top_index]+[np.nan] * (len(self.gamma_input.columns)-1)], columns=self.gamma_input.columns))
        self.gamma_input = df_top.append(self.gamma_input, ignore_index=True)
        
        # Calculating rescaled C10+ weight fractions.
        self.gamma_input['wni_lab'] = self.gamma_input['lqd_wp']/self.gamma_input['lqd_wp'].sum()
        
        # Adding a regression variable for lower C10 molecular weight boundary:
        self.gamma_input.at[0,'ubound'] = 'ita'
        self.gamma_input.iloc[-1, self.gamma_input.columns.get_loc('ubound')] = self.gamma_input.iloc[-1, self.gamma_input.columns.get_loc('ubound_init')]
        self.gamma_input['sample_id'] = id.replace('.', '_')
        self.gamma_input['heavy_end_mw'] = id.replace('.', '_')+'_heavy_mw'


# Class to store multiple sample data.
class FlashExpDataCollection(FlashExperimentData, dict):
    
    def __init__(self, worksheet_names):
        dict.__init__(self, ((wn, FlashExperimentData.__init__(self)) for wn in worksheet_names))
        
    @property
    def sample_names(self):
        return list(self.keys())
    
    def add_sample(self, sample_name, sample):
        self[sample_name] = sample
        
    def _prepare_regression(self):
        for key, item in self.items():
            item.prepare_input(key)
            # df = self._prepare_input(item.c10_heavy_end_lqd, item.av_lqd_mw, key)
            # self._collection.append(df)
        
    def gamma_distribution_fit(self, n=10, alpha=1):
        # Gamma distribution fit from C7 is not implemented yet.
        if n == 7:
            raise NotImplementedError
        self._prepare_regression()
        # reg_variables = np.concatenate(( 
        #                                 self[sample_names[0]].gamma_input.loc[self[sample_names[0]].gamma_input.index[0:-1], 'ubound'].unique())#,
                                        # self[sample_names[0]].gamma_input.loc[self._fit_df.index[0:-1], 'heavy_end_mw'].unique()))
        # reg_variables = np.delete(reg_variables, np.where(reg_variables == 100000))
        df = self[self.sample_names[0]].gamma_input
        reg_variables = np.concatenate((np.array(['alpha']), df.loc[df.index[0:-1], 'ubound'].unique(),
                                        np.array([sample.replace('.', '_')+'_heavy_mw' for sample in self.sample_names])))
        init_vals = df.loc[df.index[0:-1], 'ubound_init']
        init_vals.iloc[0] = init_vals.iloc[1]-14
        init_vals = pd.Series(alpha).append(init_vals, ignore_index=True)
        for key, item in self.items():
            init_vals = init_vals.append(pd.Series(item.ave_C10_mw), ignore_index=True)
        
        print(reg_variables)
        print(init_vals)


# Class that contains functionality to parse a Core Labs Excel report
# and pass the data to FlashExperimentData class instance.
# openpyxl only works with Excel 2010+ file format.
class CoreLabsXLSXLoader:
    
    # Depending whether a specific worksheet is input a class instance has
    # a functionality either to read the whole flash data found in the
    # workbook or just that individual worksheet.
    def __init__(self, report_path, worksheet=None):
        self.file_path = report_path
        self.worksheet = worksheet
    
    # Method to parse an individual worksheet.
    def read_flash_data(self, worksheet):
        # worksheet = ws
        # Reading the composition table as the whole.
        df = pd.read_excel(self.file_path, worksheet, skiprows = 11, nrows = 52, 
                        usecols = 'B:I', header = None,
                        names = ['scn', 'cl_name', 'lqd_mp', 'lqd_wp', 'gas_mp',
                                 'gas_wp', 'res_mp', 'res_wp'], na_values = ' ')
        # Filling in empty cells in carbon group columns.
        df['scn'] = df['scn'].ffill()
        df.set_index(['scn', 'cl_name'], inplace=True)
        # df.sort_index(inplace=True)
        liq = df.iloc[:, 0:2].reset_index()
        gas = df.iloc[:, 2:4].reset_index()
        res = df.iloc[:, 4:].reset_index()
        return liq, gas, res
    
    def read(self):
        wb = load_workbook(self.file_path)
        if self.worksheet:
            flash_data_list = self.worksheet
        else:
            flash_data_list = [worksheet for worksheet in wb.sheetnames if
                               re.search('C\.\d+', worksheet)]
        samples = FlashExpDataCollection(flash_data_list)
        for worksheet in flash_data_list:
            sheet = wb[worksheet]
            desc = sheet['B8'].value+sheet['B9'].value
            depth, sample_num, cylinder = self.__parser(desc)
            # print('Depth: ', depth, 'Sample number: ', sample_num, 'Cylinder: ', cylinder)
            liq, gas, res = self.read_flash_data(worksheet)
            # Typically, flashed liquid average mole weight in CL reports is in
            # the cell O39:
            lqd_av_mw = sheet['O39'].value
            samples.add_sample(worksheet, FlashExperimentData(liq, gas, res,
                                                              lqd_av_mw))
        return samples
    
    # Parcer function is supposed to extract useful sample descriptors from
    # the text strings above the composition table. In most Core LAbs reports
    # these strings are in the cells B8 and B9. Presumably, they contain
    # sample depth, cylinder and sample numbers. At this stage I am only
    # extracting numerical part of the these numbers. Depth is assumed to be
    # MD depth in metres.
    def __parser(self, descript_string):
        try:
            depth = float(re.search('Depth\D+(\d+\.?\d*)', descript_string).group(1))
        except:
            depth = None
        try:
            sample_num = re.search('Sample N\D+(\d+\.?\d*)', descript_string).group(1)
        except:
            sample_num = 'N/A'
        try:
            cylinder = re.search('(Cylinder|Chamber)\D+(\d+)', descript_string).group(2)
        except:
            cylinder = 'N/A'
        return depth, sample_num, cylinder

if __name__ == "__main__":
    
    path = '.\DATA\PS1.xlsx'
    cl_report = CoreLabsXLSXLoader(path, worksheet=['C.1', 'C.4'])
    sample_collection = cl_report.read()
    # print(sample_collection['C.1'].c10_heavy_end_lqd)
    sample_collection.gamma_distribution_fit()
    
    # print(sample_collection[0].c10_heavy_end_lqd)
    # cl_report.read_flash_data()